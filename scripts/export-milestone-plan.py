import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read input parameters
if len(sys.argv) < 3:
    print("Usage: python export-milestone-plan.py <input_json> <output_xlsx>")
    sys.exit(1)

input_json_path = sys.argv[1]
output_xlsx_path = sys.argv[2]

with open(input_json_path, "r", encoding="utf-8") as f:
    payload = json.load(f)

# Support both single dict payload and list payload
if isinstance(payload, list):
    plans = payload
else:
    plans = [payload]

wb = openpyxl.Workbook()
# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Helper to apply styling to range of cells
def style_range(sheet, cell_range, border=None, fill=None, font=None, alignment=None):
    for row in sheet[cell_range]:
        for cell in row:
            if border is not None:
                cell.border = border
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment

# Mappings
domain_labels = {
    "PRODUCT_ITERATION": "产品迭代",
    "PRODUCT_DELIVERY": "产品交付",
    "PRODUCT_RESEARCH": "产品调研",
    "MARKET_SUPPORT": "市场支持",
    "OPERATIONS_STABILITY": "运维稳定",
    "TECHNICAL_INNOVATION": "技术创新",
    "AI_ENABLEMENT": "AI赋能"
}

status_labels = {
    "NOT_STARTED": "未开始",
    "IN_PROGRESS": "进行中",
    "COMPLETED": "已完成",
    "DELAY_RISK": "延期风险",
    "DELAYED": "已延期",
    "PAUSED": "暂缓"
}

risk_level_labels = {
    "HIGH": "高",
    "MEDIUM": "中",
    "LOW": "低"
}

risk_status_labels = {
    "NOT_TRIGGERED": "未触发",
    "TRIGGERED": "已触发",
    "HANDLING": "处理中",
    "CLOSED": "已关闭"
}

# Fonts and Fills
title_font = Font(name="Calibri", size=10, bold=False, color="FFFFFF")
title_fill = PatternFill(fill_type="solid", start_color="0F6E56", end_color="0F6E56")

info_font = Font(name="Microsoft YaHei", size=9, color="1F3864")
info_fill = PatternFill(fill_type="solid", start_color="FFE1F5EE", end_color="FFE1F5EE")

section_title_font = Font(name="Calibri", size=10, bold=False, color="FFFFFF")
section_fill_green = PatternFill(fill_type="solid", start_color="0F6E56", end_color="0F6E56")
section_fill_red = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")

header_font = Font(name="Noto Sans CJK SC", size=9, bold=True, color="FFFFFF")
header_fill_blue = PatternFill(fill_type="solid", start_color="1F3864", end_color="1F3864")
header_fill_red = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")

goal_row_fill = PatternFill(fill_type="solid", start_color="FFE1F5EE", end_color="FFE1F5EE")
risk_row_fill_even = PatternFill(fill_type="solid", start_color="FFFCE4EC", end_color="FFFCE4EC")
risk_row_fill_odd = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")

# Borders
thin_side = Side(style='thin', color='BFBFBF')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

align_center_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Loop and build worksheets
for data in plans:
    team_name = data.get("teamName", "项目组")
    sheet = wb.create_sheet(title=team_name)
    
    # Ensure gridlines are visible
    sheet.views.sheetView[0].showGridLines = True
    
    # Define Column Widths
    widths = {
        "A": 2.6,
        "B": 5.6,
        "C": 13.6,
        "D": 22.6,
        "E": 24.6,
        "F": 13.6,
        "G": 10.6,
        "H": 17.1,
        "I": 8.6,
        "J": 18.5,
        "K": 8.6,
        "L": 12.2,
        "M": 8.6,
        "N": 13.0,
        "O": 13.0,
        "P": 13.0
    }
    for col, w in widths.items():
        sheet.column_dimensions[col].width = w
    
    # Row heights
    sheet.row_dimensions[1].height = 6.0
    sheet.row_dimensions[2].height = 39.0
    sheet.row_dimensions[3].height = 7.5
    sheet.row_dimensions[4].height = 21.0
    sheet.row_dimensions[5].height = 15.0
    sheet.row_dimensions[6].height = 27.0
    sheet.row_dimensions[7].height = 24.75
    
    # 1. Row 2: Sheet Title
    products = data.get("products", [])
    products_str = "/".join(products) if products else ""
    title_text = f"技术中心 · {team_name}  季度里程碑计划表"
    if products_str:
        title_text += f"  |  覆盖产品：{products_str}"
    
    sheet["B2"] = title_text
    sheet.merge_cells("B2:P2")
    style_range(sheet, "B2:P2", border=thin_border, fill=title_fill, font=title_font, alignment=align_center_center)
    
    # 2. Row 4: Header Info
    quarter_months = {1: "1-3月", 2: "4-6月", 3: "7-9月", 4: "10-12月"}.get(data.get("quarter", 3), "7-9月")
    sheet["B4"] = f"季度: Q{data.get('quarter', 3)}({data.get('year', 2026)}年{quarter_months})"
    sheet.merge_cells("B4:E4")
    
    sheet["F4"] = f"产品线组: {team_name}"
    sheet.merge_cells("F4:I4")
    
    sheet["J4"] = f"组长: {data.get('leaderName', '未设置')}"
    sheet.merge_cells("J4:M4")
    
    sheet["N4"] = "版本: V1.0"
    sheet.merge_cells("N4:P4")
    
    style_range(sheet, "B4:E4", border=thin_border, fill=info_fill, font=info_font, alignment=align_left_center)
    style_range(sheet, "F4:I4", border=thin_border, fill=info_fill, font=info_font, alignment=align_left_center)
    style_range(sheet, "J4:M4", border=thin_border, fill=info_fill, font=info_font, alignment=align_left_center)
    style_range(sheet, "N4:P4", border=thin_border, fill=info_fill, font=info_font, alignment=align_left_center)
    
    # 3. Row 6: Section A Title
    sheet["B6"] = "A  季度产品目标看板"
    sheet.merge_cells("B6:P6")
    style_range(sheet, "B6:P6", border=thin_border, fill=section_fill_green, font=section_title_font, alignment=align_left_center)
    
    # 4. Row 7: Goals Headers
    headers_goals = ["序", "目标域", "季度目标", "成功标准", "M1目标", "M1状态", "M2目标", "M2状态", "M3目标", "M3状态", "当前完成", "达成率", "季度状态", "关键依赖", "备注"]
    for idx, h in enumerate(headers_goals):
        col_letter = get_column_letter(2 + idx)
        sheet[f"{col_letter}7"] = h
    style_range(sheet, "B7:P7", border=thin_border, fill=header_fill_blue, font=header_font, alignment=align_center_center)
    
    # 5. Row 8+: Goals Data
    current_row = 8
    
    goals_data = data.get("goals", [])
    for index, goal in enumerate(goals_data):
        sheet.row_dimensions[current_row].height = None # Auto fit height
    
        sheet.cell(row=current_row, column=2, value=index + 1)
        sheet.cell(row=current_row, column=3, value=domain_labels.get(goal.get('domain'), ''))
        sheet.cell(row=current_row, column=4, value=goal.get('quarterlyGoal'))
        sheet.cell(row=current_row, column=5, value=goal.get('successCriteria'))
        sheet.cell(row=current_row, column=6, value=goal.get('month1Goal'))
        sheet.cell(row=current_row, column=7, value=status_labels.get(goal.get('month1Status'), ''))
        sheet.cell(row=current_row, column=8, value=goal.get('month2Goal'))
        sheet.cell(row=current_row, column=9, value=status_labels.get(goal.get('month2Status'), ''))
        sheet.cell(row=current_row, column=10, value=goal.get('month3Goal'))
        sheet.cell(row=current_row, column=11, value=status_labels.get(goal.get('month3Status'), ''))
        sheet.cell(row=current_row, column=12, value=goal.get('currentCompletion'))
        rate = goal.get('achievementRate', 0)
        sheet.cell(row=current_row, column=13, value=f"{rate}%" if rate is not None else "")
        sheet.cell(row=current_row, column=14, value=status_labels.get(goal.get('quarterlyStatus'), ''))
        sheet.cell(row=current_row, column=15, value=goal.get('keyDependencies'))
        sheet.cell(row=current_row, column=16, value=goal.get('notes'))
        
        # Fonts
        sheet.cell(row=current_row, column=2).font = Font(name="Arial", size=10, bold=True, color="0F6E56")
        sheet.cell(row=current_row, column=3).font = Font(name="宋体", size=10, bold=True, color="1F3864")
        for col_idx in range(4, 17):
            sheet.cell(row=current_row, column=col_idx).font = Font(name="宋体", size=10, color="000000")
            
        # Alignments
        sheet.cell(row=current_row, column=2).alignment = align_center_center
        sheet.cell(row=current_row, column=3).alignment = align_left_center
        for col_idx in range(4, 7):
            sheet.cell(row=current_row, column=col_idx).alignment = align_left_center
        sheet.cell(row=current_row, column=7).alignment = align_center_center
        sheet.cell(row=current_row, column=8).alignment = align_left_center
        sheet.cell(row=current_row, column=9).alignment = align_center_center
        sheet.cell(row=current_row, column=10).alignment = align_left_center
        sheet.cell(row=current_row, column=11).alignment = align_center_center
        sheet.cell(row=current_row, column=12).alignment = align_left_center
        sheet.cell(row=current_row, column=13).alignment = align_center_center
        sheet.cell(row=current_row, column=14).alignment = align_center_center
        sheet.cell(row=current_row, column=15).alignment = align_left_center
        sheet.cell(row=current_row, column=16).alignment = align_left_center
        
        # Fill & Border
        for col_idx in range(2, 17):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.fill = goal_row_fill
            cell.border = thin_border
            
        current_row += 1
    
    # Separator row
    sheet.row_dimensions[current_row].height = 36.0
    for col_idx in range(2, 17):
        sheet.cell(row=current_row, column=col_idx).border = Border() # No border
    current_row += 1
    
    # Section B Title
    sheet.row_dimensions[current_row].height = 36.0
    sheet.cell(row=current_row, column=2, value="B  季度风险登记册")
    sheet.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=16)
    style_range(sheet, f"B{current_row}:P{current_row}", border=thin_border, fill=section_fill_red, font=section_title_font, alignment=align_left_center)
    current_row += 1
    
    # Risks Header
    sheet.row_dimensions[current_row].height = 36.0
    headers_risks = ["风险编号", "风险描述", "影响里程碑", "概率", "影响", "综合级", "触发条件", "应对策略", "预警节点", "当前状态"]
    for idx, h in enumerate(headers_risks):
        sheet.cell(row=current_row, column=2 + idx, value=h)
    style_range(sheet, f"B{current_row}:K{current_row}", border=thin_border, fill=header_fill_red, font=header_font, alignment=align_center_center)
    # Clear extra columns on risk header row
    for col_idx in range(12, 17):
        sheet.cell(row=current_row, column=col_idx).border = Border()
    current_row += 1
    
    # Risks Data
    risks_data = data.get("risks", [])
    if len(risks_data) == 0:
        # Render one empty styled risk row
        sheet.row_dimensions[current_row].height = 36.0
        fill = risk_row_fill_even
        for col_idx in range(2, 12):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = align_center_center if col_idx in [2, 5, 6, 7, 11] else align_left_center
        # Sequence empty
        sheet.cell(row=current_row, column=2, value="")
        sheet.cell(row=current_row, column=2).font = Font(name="Arial", size=10, bold=True, color="C00000")
        for col_idx in range(3, 12):
            sheet.cell(row=current_row, column=col_idx).font = Font(name="Noto Sans CJK SC", size=10, color="000000")
        current_row += 1
    else:
        for index, risk in enumerate(risks_data):
            sheet.row_dimensions[current_row].height = 36.0
            fill = risk_row_fill_even if index % 2 == 0 else risk_row_fill_odd
            
            sheet.cell(row=current_row, column=2, value=f"R{index + 1}")
            sheet.cell(row=current_row, column=3, value=risk.get('riskDescription'))
            sheet.cell(row=current_row, column=4, value=risk.get('affectedMilestone'))
            sheet.cell(row=current_row, column=5, value=risk_level_labels.get(risk.get('probability'), ''))
            sheet.cell(row=current_row, column=6, value=risk_level_labels.get(risk.get('impact'), ''))
            sheet.cell(row=current_row, column=7, value=risk_level_labels.get(risk.get('overallLevel'), ''))
            sheet.cell(row=current_row, column=8, value=risk.get('triggerCondition'))
            sheet.cell(row=current_row, column=9, value=risk.get('responseStrategy'))
            sheet.cell(row=current_row, column=10, value=risk.get('warningPoint'))
            sheet.cell(row=current_row, column=11, value=risk_status_labels.get(risk.get('status'), ''))
            
            # Fonts
            sheet.cell(row=current_row, column=2).font = Font(name="Arial", size=10, bold=True, color="C00000")
            for col_idx in range(3, 12):
                sheet.cell(row=current_row, column=col_idx).font = Font(name="Noto Sans CJK SC", size=10, color="000000")
                
            # Alignments & Styles
            for col_idx in range(2, 12):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = align_center_center if col_idx in [2, 5, 6, 7, 11] else align_left_center
                
            current_row += 1

# Save Workbook
wb.save(output_xlsx_path)
print("Successfully generated Excel milestone plan.")
