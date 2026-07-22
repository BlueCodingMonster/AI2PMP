import openpyxl
import re
import json
import os
import datetime

path = r"d:\workspace-ai\AI2PmP\doc\技术中心月度（7月）项目经营计划.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

def clean(v):
    if v is None:
        return ""
    return str(v).strip()

def format_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%dT00:00:00.000Z")
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        parts = re.split(r"[-/. ]", val_str)
        if len(parts) >= 3:
            y = int(parts[0])
            m = int(parts[1])
            d = int(parts[2])
            return f"{y:04d}-{m:02d}-{d:02d}T00:00:00.000Z"
    except Exception:
        pass
    return None

def map_risk_level(val):
    val_str = clean(val)
    if "高" in val_str:
        return "HIGH"
    if "中" in val_str:
        return "MEDIUM"
    if "低" in val_str:
        return "LOW"
    return None

def map_request_type(val):
    val_str = clean(val)
    if "人员" in val_str:
        return "PEOPLE"
    if "技术" in val_str:
        return "TECHNOLOGY"
    if "预算" in val_str:
        return "BUDGET"
    if "管理" in val_str or "协调" in val_str:
        return "MANAGEMENT_COORDINATION"
    return None

sections = [
    "一、产品交付计划",
    "二、项目交付计划",
    "三、市场化动作计划",
    "四、成本优化计划",
    "五、AI+赋能产品计划",
    "六、AI+效能提升计划",
    "七、项目风险",
    "八、资源需求"
]

all_plans = []

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Parsing sheet: {sheet_name}")
    
    # Parse header info from Row 1
    # D1 is usually the month (e.g. "7月")
    month_val = sheet.cell(row=1, column=4).value
    group_leader = sheet.cell(row=1, column=3).value
    
    month_match = re.search(r"(\d+)月", clean(month_val))
    if not month_match:
        print(f"  Warning: Could not parse month from '{month_val}' in sheet '{sheet_name}'. Skipping.")
        continue
    month = int(month_match.group(1))
    year = 2026  # default based on cell dates
    
    team_match = re.search(r"小组与组长[：:]([^/]+)", clean(group_leader))
    if not team_match:
        # Fallback to sheet_name
        team_name = sheet_name.strip()
    else:
        team_name = team_match.group(1).strip()
        
    print(f"  Detected team: '{team_name}', year: {year}, month: {month}")
    
    product_deliveries = []
    project_deliveries = []
    market_actions = []
    cost_optimizations = []
    ai_product_enablements = []
    ai_efficiencies = []
    risks = []
    resource_requests = []
    
    current_section = None
    
    # Read rows sequentially up to row 100
    for r in range(1, 100):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        val_a = clean(row_vals[0])
        
        # Check if this row is a section header
        found_section = False
        for sec in sections:
            if sec in val_a:
                current_section = sec
                found_section = True
                break
                
        if found_section:
            continue
            
        if current_section:
            # Check if row is empty
            non_empty_str = [clean(v) for v in row_vals if v is not None and clean(v)]
            if not non_empty_str:
                continue
                
            # Check if it is the section's table header, skip
            if val_a in ["模块/功能/版本", "项目名称", "产品/项目", "优化项", "事项", "风险项", "需求类型"]:
                continue
                
            # Parse row based on current section
            if current_section == "一、产品交付计划":
                # Col 1: moduleVersion, Col 2: deliveryContent, Col 3: plannedCompletionDate
                module_ver = clean(row_vals[0])
                content = clean(row_vals[1])
                date_str = format_date(row_vals[2])
                if module_ver or content or date_str:
                    product_deliveries.append({
                        "moduleVersion": module_ver or None,
                        "deliveryContent": content or None,
                        "plannedCompletionDate": date_str
                    })
                    
            elif current_section == "二、项目交付计划":
                # Col 1: projectName, Col 2: deliveryContent, Col 3: plannedCompletionDate, Col 4: customerName
                proj_name = clean(row_vals[0])
                content = clean(row_vals[1])
                date_str = format_date(row_vals[2])
                cust_name = clean(row_vals[3])
                if proj_name or content or date_str or cust_name:
                    project_deliveries.append({
                        "projectName": proj_name or None,
                        "deliveryContent": content or None,
                        "plannedCompletionDate": date_str,
                        "customerName": cust_name or None
                    })
                    
            elif current_section == "三、市场化动作计划":
                # Col 1: productOrProject, Col 2: marketAction, Col 3: outputResult, Col 4: plannedCompletionDate
                prod_proj = clean(row_vals[0])
                action = clean(row_vals[1])
                result = clean(row_vals[2])
                date_str = format_date(row_vals[3])
                if prod_proj or action or result or date_str:
                    market_actions.append({
                        "productOrProject": prod_proj or None,
                        "marketAction": action or None,
                        "outputResult": result or None,
                        "plannedCompletionDate": date_str
                    })
                    
            elif current_section == "四、成本优化计划":
                # Col 1: optimizationItem, Col 2: currentProblem, Col 3: optimizationMeasure
                item = clean(row_vals[0])
                prob = clean(row_vals[1])
                meas = clean(row_vals[2])
                # Skip if both problem and measure are empty
                if prob or meas:
                    cost_optimizations.append({
                        "optimizationItem": item or None,
                        "currentProblem": prob or None,
                        "optimizationMeasure": meas or None
                    })
                    
            elif current_section == "五、AI+赋能产品计划":
                # Col 1: item, Col 2: outputResult, Col 3: plannedCompletionDate
                item = clean(row_vals[0])
                result = clean(row_vals[1])
                date_str = format_date(row_vals[2])
                if item or result or date_str:
                    ai_product_enablements.append({
                        "item": item or None,
                        "outputResult": result or None,
                        "plannedCompletionDate": date_str
                    })
                    
            elif current_section == "六、AI+效能提升计划":
                # Col 1: item, Col 2: outputResult, Col 3: plannedCompletionDate
                item = clean(row_vals[0])
                result = clean(row_vals[1])
                date_str = format_date(row_vals[2])
                if item or result or date_str:
                    ai_efficiencies.append({
                        "item": item or None,
                        "outputResult": result or None,
                        "plannedCompletionDate": date_str
                    })
                    
            elif current_section == "七、项目风险":
                # Col 1: riskItem, Col 2: riskLevel, Col 3: impactScope, Col 4: responseMeasure
                item = clean(row_vals[0])
                level = map_risk_level(row_vals[1])
                scope = clean(row_vals[2])
                measure = clean(row_vals[3])
                # Skip if all fields except level are empty (empty template row)
                if item or scope or measure:
                    risks.append({
                        "riskItem": item or None,
                        "riskLevel": level,
                        "impactScope": scope or None,
                        "responseMeasure": measure or None
                    })
                    
            elif current_section == "八、资源需求":
                # Col 1: requestType, Col 2: content, Col 3: urgency, Col 4: supportDepartment
                req_type = map_request_type(row_vals[0])
                content = clean(row_vals[1])
                urg = map_risk_level(row_vals[2])
                dept = clean(row_vals[3])
                # Skip if content is empty (no request made)
                if content:
                    resource_requests.append({
                        "requestType": req_type,
                        "content": content or None,
                        "urgency": urg,
                        "supportDepartment": dept or None
                    })
                    
    all_plans.append({
        "teamName": team_name,
        "year": year,
        "month": month,
        "productDeliveries": product_deliveries,
        "projectDeliveries": project_deliveries,
        "marketActions": market_actions,
        "costOptimizations": cost_optimizations,
        "aiProductEnablements": ai_product_enablements,
        "aiEfficiencies": ai_efficiencies,
        "risks": risks,
        "resourceRequests": resource_requests
    })

# Write output JSON
out_path = r"d:\workspace-ai\AI2PmP\scripts\monthly-plan.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(all_plans, f, ensure_ascii=False, indent=2)

print(f"\nSuccessfully parsed {len(all_plans)} sheets and wrote JSON to {out_path}")
