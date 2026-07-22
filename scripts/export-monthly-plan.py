import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import sys
import os

def clean(v):
    if v is None:
        return ""
    return str(v).strip()

def format_date(val):
    if not val:
        return ""
    # Usually "2026-07-31T00:00:00.000Z" -> slice to "2026-07-31"
    val_str = str(val).strip()
    if len(val_str) >= 10 and val_str[4] == '-' and val_str[7] == '-':
        return val_str[:10]
    return val_str

def map_risk_level_to_zh(level):
    if not level:
        return ""
    mapping = {
        "HIGH": "高",
        "MEDIUM": "中",
        "LOW": "低"
    }
    return mapping.get(clean(level), "")

def style_cell(cell, font_name="Songti SC", font_size=11, bold=False, italic=False, font_color=None,
               fill_color=None, alignment_h=None, alignment_v=None, wrap_text=False, border=None):
    if font_name or font_size:
        cell.font = Font(name=font_name, size=font_size, bold=bold, italic=italic, color=font_color)
    if fill_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
    if alignment_h or alignment_v or wrap_text:
        cell.alignment = Alignment(horizontal=alignment_h, vertical=alignment_v, wrap_text=wrap_text)
    if border:
        cell.border = border

def main():
    if len(sys.argv) < 3:
        print("Usage: python export-monthly-plan.py <payload_json_path> <output_xlsx_path>")
        sys.exit(1)

    payload_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(payload_path, "r", encoding="utf-8") as f:
        plans = json.load(f)

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for plan in plans:
        team_name = plan.get("teamName", "未知小组")
        leader_name = plan.get("leaderName", "—")
        year = plan.get("year", 2026)
        month = plan.get("month", 7)

        ws = wb.create_sheet(title=team_name[:30])
        ws.views.sheetView[0].showGridLines = True

        # Column widths
        col_widths = {
            "A": 45,
            "B": 48,
            "C": 50,
            "D": 27,
            "E": 14,
            "F": 56,
            "G": 14
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Row 1 Height
        ws.row_dimensions[1].height = 54.0

        # Title block fills (A1 to E1 are green)
        for col_idx in range(1, 6):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            cell = ws[f"{col_letter}1"]
            style_cell(cell, font_name="Songti SC", font_size=12, bold=True, fill_color="FF98D7B6", alignment_v="center")

        # Merge A1:B1 and set title
        ws.merge_cells("A1:B1")
        ws["A1"].value = "技术中心月度项目经营计划   "
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # C1 Group and Leader
        ws["C1"].value = f"小组与组长：{team_name}/{leader_name}"
        ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

        # D1 Month
        ws["D1"].value = f"{month}月"
        ws["D1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["D1"].font = Font(name="Songti SC", size=12, bold=False)

        # Column F instructions (F1:F14 merged)
        ws.merge_cells("F1:F14")
        instruction_text = (
            "说明：\n\n"
            "1、提交时间：每月结束前5天提交至部门经理处；\n"
            "2、计划说明：务必结合季度里程碑目标进行当月计划拆解和填写；"
        )
        ws["F1"].value = instruction_text
        
        # Style all cells in merged range F1:F14
        for r_idx in range(1, 15):
            cell = ws.cell(row=r_idx, column=6) # Column F is 6
            style_cell(cell, font_name="Songti SC", font_size=11, bold=False, fill_color="FFC5CAD3",
                       alignment_h="left", alignment_v="top", wrap_text=True, border=thin_border)

        sections_def = [
            {
                "title": "一、产品交付计划",
                "headers": ["模块/功能/版本", "交付内容（必须有书面文件记录）", "计划完成时间"],
                "data_key": "productDeliveries",
                "col_count": 3,
                "fields": ["moduleVersion", "deliveryContent", "plannedCompletionDate"]
            },
            {
                "title": "二、项目交付计划",
                "headers": ["项目名称", "交付内容（必须有书面文件记录）", "计划完成时间", "客户名称"],
                "data_key": "projectDeliveries",
                "col_count": 4,
                "fields": ["projectName", "deliveryContent", "plannedCompletionDate", "customerName"]
            },
            {
                "title": "三、市场化动作计划",
                "headers": ["产品/项目", "市场化动作", "输出成果（必须有书面文件记录）", "计划完成时间"],
                "data_key": "marketActions",
                "col_count": 4,
                "fields": ["productOrProject", "marketAction", "outputResult", "plannedCompletionDate"]
            },
            {
                "title": "四、成本优化计划",
                "headers": ["优化项", "当前问题", "优化措施"],
                "data_key": "costOptimizations",
                "col_count": 3,
                "fields": ["optimizationItem", "currentProblem", "optimizationMeasure"],
                "fixed_items": ["云资源", "第三方服务", "开发效率", "测试效率", "人员优化"]
            },
            {
                "title": "五、AI+赋能产品计划",
                "headers": ["事项", "输出成果", "计划完成时间"],
                "data_key": "aiProductEnablements",
                "col_count": 3,
                "fields": ["item", "outputResult", "plannedCompletionDate"]
            },
            {
                "title": "六、AI+效能提升计划",
                "headers": ["事项", "输出成果", "计划完成时间"],
                "data_key": "aiEfficiencies",
                "col_count": 3,
                "fields": ["item", "outputResult", "plannedCompletionDate"]
            },
            {
                "title": "七、项目风险",
                "headers": ["风险项", "风险等级", "影响范围", "应对措施"],
                "data_key": "risks",
                "col_count": 4,
                "fields": ["riskItem", "riskLevel", "impactScope", "responseMeasure"]
            },
            {
                "title": "八、资源需求",
                "headers": ["需求类型", "具体内容", "紧急程度", "需要支持部门/小组"],
                "data_key": "resourceRequests",
                "col_count": 4,
                "fields": ["requestType", "content", "urgency", "supportDepartment"],
                "fixed_items": ["人员", "技术", "预算", "管理协调"]
            }
        ]

        r = 2
        for sec in sections_def:
            # 1. Section Title Row
            ws.row_dimensions[r].height = 20.0
            for col_idx in range(1, 5):
                cell = ws.cell(row=r, column=col_idx)
                style_cell(cell, font_name="Songti SC", font_size=11, bold=True, fill_color="FFC3EAD5",
                           alignment_v="center", wrap_text=True, border=thin_border)
            ws.cell(row=r, column=1).value = sec["title"]
            r += 1

            # 2. Table Header Row
            ws.row_dimensions[r].height = 18.0
            for col_idx, header in enumerate(sec["headers"], 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.value = header
                style_cell(cell, font_name="Songti SC", font_size=11, bold=False, fill_color="FFC5CAD3",
                           alignment_v="top", wrap_text=True, border=thin_border)
            # Ensure empty header cells in title block (up to 4 cols) get styled correctly if needed?
            # No, only columns that actually have headers are filled.
            r += 1

            # 3. Data Rows
            data_items = plan.get(sec["data_key"], [])
            col_count = sec["col_count"]

            if "fixed_items" in sec:
                # Fixed items for Cost Optimizations (Section 4) and Resource Requests (Section 8)
                fixed_list = sec["fixed_items"]
                for item_name in fixed_list:
                    ws.row_dimensions[r].height = 22.0
                    
                    matching_data = None
                    if sec["data_key"] == "costOptimizations":
                        for d_item in data_items:
                            if clean(d_item.get("optimizationItem")) == item_name:
                                matching_data = d_item
                                break
                    elif sec["data_key"] == "resourceRequests":
                        # map DB enum back to Chinese fixed item
                        type_mapping = {
                            "PEOPLE": "人员",
                            "TECHNOLOGY": "技术",
                            "BUDGET": "预算",
                            "MANAGEMENT_COORDINATION": "管理协调"
                        }
                        for d_item in data_items:
                            db_type = clean(d_item.get("requestType"))
                            if type_mapping.get(db_type) == item_name:
                                matching_data = d_item
                                break

                    # Style columns A to col_count
                    for c_idx in range(1, col_count + 1):
                        cell = ws.cell(row=r, column=c_idx)
                        style_cell(cell, font_name="Songti SC", font_size=11, bold=False,
                                   alignment_v="top", wrap_text=True, border=thin_border)

                    ws.cell(row=r, column=1).value = item_name
                    
                    if matching_data:
                        if sec["data_key"] == "costOptimizations":
                            ws.cell(row=r, column=2).value = clean(matching_data.get("currentProblem"))
                            ws.cell(row=r, column=3).value = clean(matching_data.get("optimizationMeasure"))
                        elif sec["data_key"] == "resourceRequests":
                            ws.cell(row=r, column=2).value = clean(matching_data.get("content"))
                            ws.cell(row=r, column=3).value = map_risk_level_to_zh(matching_data.get("urgency"))
                            ws.cell(row=r, column=4).value = clean(matching_data.get("supportDepartment"))
                    
                    r += 1
            else:
                # Dynamic list sections (1, 2, 3, 5, 6, 7)
                if not data_items:
                    # Write exactly one empty row
                    ws.row_dimensions[r].height = 22.0
                    for c_idx in range(1, col_count + 1):
                        cell = ws.cell(row=r, column=c_idx)
                        style_cell(cell, font_name="Songti SC", font_size=11, bold=False,
                                   alignment_v="top", wrap_text=True, border=thin_border)
                    
                    # Risk (Section 7) empty row has a placeholder "高/中/低" in column B
                    if sec["data_key"] == "risks":
                        ws.cell(row=r, column=2).value = "高/中/低"
                        
                    r += 1
                else:
                    for item in data_items:
                        ws.row_dimensions[r].height = 22.0
                        for c_idx, field in enumerate(sec["fields"], 1):
                            cell = ws.cell(row=r, column=c_idx)
                            style_cell(cell, font_name="Songti SC", font_size=11, bold=False,
                                       alignment_v="top", wrap_text=True, border=thin_border)
                            
                            val = item.get(field)
                            if field == "plannedCompletionDate":
                                val = format_date(val)
                            elif field == "riskLevel":
                                val = map_risk_level_to_zh(val)
                            else:
                                val = clean(val)
                                
                            cell.value = val
                        r += 1

            # 4. Spacing row between sections
            ws.row_dimensions[r].height = 16.5
            r += 1

    wb.save(output_path)
    print(f"Successfully generated monthly Excel file at {output_path}")

if __name__ == "__main__":
    main()
